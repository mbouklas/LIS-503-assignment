import openpyxl

iStudent_start_row = 3
iStudent_column = 3
iAssignment_name_row = 1
iAssignment_start_column = 6
iAssignment_category_row = 2

iWeighting_name_row = 1
iWeighting_value_row = 2
iWeighting_start_column = 1

sPath = "studentdata.xlsx"

sWorkbook = openpyxl.load_workbook(sPath)

sGrade_sheet = sWorkbook["Grades"]
sWeight_sheet = sWorkbook["Weighting"]

def get_student_names():
    lNames = []
    iCurr_row = iStudent_start_row
    
    iCell = sGrade_sheet.cell(row = iCurr_row, column = iStudent_column)
    while iCell.value is not None:
        lNames.append(iCell.value)
        iCurr_row = iCurr_row + 1
        iCell = sGrade_sheet.cell(row = iCurr_row, column = iStudent_column)
    return lNames
        
def get_grades_by_name(sName):
    iCurr_row = iStudent_start_row
    
    iCell = sGrade_sheet.cell(row = iCurr_row, column = iStudent_column)
    while iCell.value != sName:
        iCurr_row = iCurr_row + 1
        iCell = sGrade_sheet.cell(row = iCurr_row, column = iStudent_column)

    iCurr_column = iAssignment_start_column
    lGrades = []
    iCell = sGrade_sheet.cell (row = iCurr_row, column = iCurr_column)
    while iCell.value is not None:
        sCategory = sGrade_sheet.cell(row = iAssignment_category_row, column = iCurr_column)
        sAssignment_name = sGrade_sheet.cell(row = iAssignment_name_row, column = iCurr_column)

        assignment_dict = {
            "Name": sAssignment_name.value,
            "Category": sCategory.value,
            "Grade": iCell.value
            }
        lGrades.append(assignment_dict)
        iCurr_column = iCurr_column + 1
        
        iCell = sGrade_sheet.cell(row = iCurr_row, column = iCurr_column)
        
    return lGrades
        
def load_weightings():
    weightings_dict = {}
    iCurr_column = iWeighting_start_column
    sCat_name = sWeight_sheet.cell(row = iWeighting_name_row, column = iCurr_column)
    while sCat_name.value is not None:
        sCat_value = sWeight_sheet.cell(row = iWeighting_value_row, column = iCurr_column)
        weightings_dict[sCat_name.value] = sCat_value.value
        iCurr_column = iCurr_column + 1 
        sCat_name = sWeight_sheet.cell(row = iWeighting_name_row, column = iCurr_column)
    return weightings_dict

print(load_weightings())


    
